#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 Motor_Para.xlsx 的四个表格中提取带有变量名的参数，生成 C 数组定义文件。

用法: python generate_lookup_tables.py [输出文件路径]
"""

import openpyxl
import os
import sys
import re

# Unicode 码点: '变量名'
_MARKER_VAR_NAME_CODEPOINTS = (0x53d8, 0x91cf, 0x540d)

# C 变量名正则
_RE_C_IDENTIFIER = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')


def _is_var_name_marker(cell_value):
    """检查单元格是否为'变量名'标记（用 Unicode 码点比较，避免编码问题）。"""
    if not isinstance(cell_value, str):
        return False
    if len(cell_value) != len(_MARKER_VAR_NAME_CODEPOINTS):
        return False
    return all(ord(c) == cp for c, cp in zip(cell_value, _MARKER_VAR_NAME_CODEPOINTS))


def fmt_float(v):
    """将数值格式化为 C float 字面量。"""
    if isinstance(v, int):
        return f"{v}.0f"
    s = f"{v:.8g}"
    if 'e' in s or 'E' in s:
        return s + 'f'
    if '.' in s:
        return s + 'f'
    return s + '.0f'


def _try_extract_2d_table(ws, var_row):
    """尝试从工作表中提取二维查表数组。

    预期的布局格式：
      var_row:     列A 为 '变量名' 标记
      var_row + 1: 列A 为 C 标识符（二维表变量名）
      var_row + 2: 列B~... 为列索引头（数值，1..N）
      var_row + 3+: 列A 为行索引（数值），列B~... 为数据（N列）

    返回 (var_name, 2d_values) 或 None。
    """
    # 检查 (var_row + 1, 列A) 是否为 C 标识符
    name_cell = ws.cell(row=var_row + 1, column=1).value
    if not isinstance(name_cell, str) or not _RE_C_IDENTIFIER.match(name_cell.strip()):
        return None

    var_name = name_cell.strip()

    # 从 var_row + 2 的列B开始读取列数
    col_count = 0
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=var_row + 2, column=col).value
        if isinstance(val, (int, float)):
            col_count += 1
        else:
            break

    if col_count == 0:
        return None

    # 从 var_row + 3 开始读取数据行
    rows_data = []
    for row_idx in range(var_row + 3, ws.max_row + 1):
        row_header = ws.cell(row=row_idx, column=1).value
        if not isinstance(row_header, (int, float)):
            break

        row_values = []
        for col in range(2, 2 + col_count):
            val = ws.cell(row=row_idx, column=col).value
            row_values.append(float(val) if isinstance(val, (int, float)) else 0.0)

        if len(row_values) == col_count:
            rows_data.append(row_values)
        else:
            break

    if not rows_data:
        return None

    return var_name, rows_data


def extract_tables(xlsx_path):
    """从 Excel 中提取所有带变量名的查表数组。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_tables = {}  # var_name -> [float_values]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 方案1: 查找"变量名"标记（可能在列A或列B）
        var_row = None
        for row_idx in range(1, ws.max_row + 1):
            if _is_var_name_marker(ws.cell(row=row_idx, column=1).value) or \
               _is_var_name_marker(ws.cell(row=row_idx, column=2).value):
                var_row = row_idx
                break

        # 方案2 (备选): 从下往上找第一个含C标识符且其上方有数据的行
        if var_row is None:
            for row_idx in range(ws.max_row, 0, -1):
                # 检查该行是否在 列2~max_column 范围内有C标识符
                has_identifier = False
                for col in range(2, ws.max_column + 1):
                    cell_val = ws.cell(row=row_idx, column=col).value
                    if isinstance(cell_val, str) and _RE_C_IDENTIFIER.match(cell_val.strip()):
                        has_identifier = True
                        break
                if not has_identifier:
                    continue
                # 检查上方行是否有数字数据
                for check_r in range(row_idx - 1, max(row_idx - 6, 0), -1):
                    for col in range(2, ws.max_column + 1):
                        if isinstance(ws.cell(row=check_r, column=col).value, (int, float)):
                            var_row = row_idx
                            break
                    if var_row:
                        break
                if var_row:
                    break

        if var_row is None:
            print(f"  Skip '{sheet_name}': var-name row not found")
            continue

        # 收集变量定义：列号 -> 变量名 (过滤掉非C标识符的单元格)
        var_defs = {}
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=var_row, column=col)
            if isinstance(cell.value, str) and _RE_C_IDENTIFIER.match(cell.value.strip()):
                var_defs[col] = cell.value.strip()

        if not var_defs:
            # 尝试二维表提取（变量名在列A下方，数据在右侧网格中）
            result_2d = _try_extract_2d_table(ws, var_row)
            if result_2d:
                var_name, values_2d = result_2d
                if var_name in all_tables:
                    print(f"    Warning: '{var_name}' redefined, overwriting")
                all_tables[var_name] = values_2d
                print(f"  {sheet_name}: 2D table '{var_name}' [{len(values_2d)}][{len(values_2d[0])}]")
            else:
                print(f"  Skip '{sheet_name}': no variable names in row {var_row}")
            continue

        # 查找数据行：从 var_row-1 往上，跳过紧邻的空行（间距行），
        # 然后收集连续的数据行直到遇到非数据行。
        data_rows = []
        gap_allowed = 3
        gap_count = 0
        in_data = False
        for row_idx in range(var_row - 1, 0, -1):
            has_data = any(
                isinstance(ws.cell(row=row_idx, column=col).value, (int, float))
                for col in var_defs
            )
            if has_data:
                data_rows.insert(0, row_idx)
                in_data = True
                gap_count = 0
            elif in_data:
                break
            else:
                gap_count += 1
                if gap_count > gap_allowed:
                    break

        if not data_rows:
            # 也尝试二维表（可能变量名行正下方就是数据）
            result_2d = _try_extract_2d_table(ws, var_row)
            if result_2d:
                var_name, values_2d = result_2d
                if var_name in all_tables:
                    print(f"    Warning: '{var_name}' redefined, overwriting")
                all_tables[var_name] = values_2d
                print(f"  {sheet_name}: 2D table '{var_name}' [{len(values_2d)}][{len(values_2d[0])}]")
            else:
                print(f"  Skip '{sheet_name}': no data rows found")
            continue

        print(f"  {sheet_name}: {len(data_rows)} rows, {len(var_defs)} vars")

        # 提取每个变量的数据
        for col, var_name in var_defs.items():
            values = []
            for row_idx in data_rows:
                val = ws.cell(row=row_idx, column=col).value
                values.append(float(val) if isinstance(val, (int, float)) else 0.0)

            if var_name in all_tables:
                print(f"    Warning: '{var_name}' redefined, overwriting")

            all_tables[var_name] = values

    wb.close()
    return all_tables


def generate_c(tables, output_path):
    """生成 .c 文件。"""
    lines = [
        '// Auto-generated by generate_lookup_tables.py from Motor_Para.xlsx',
        '// DO NOT EDIT manually.',
        '',
        '#ifdef MOTOR_CONFIG_H',
        '',
    ]

    for var_name in sorted(tables.keys()):
        values = tables[var_name]

        if values and isinstance(values[0], list):
            # 二维数组
            rows = len(values)
            cols = len(values[0])
            formatted_rows = []
            for row in values:
                formatted_vals = ', '.join(fmt_float(v) for v in row)
                formatted_rows.append(f'    {{{formatted_vals}}},')

            lines.append(f'const float {var_name}[{rows}][{cols}] = {{')
            lines.extend(formatted_rows)
            lines.append('};')
        else:
            # 一维数组
            count = len(values)

            # 每行 6 个元素，每行末尾加逗号
            formatted_lines = []
            for i in range(0, len(values), 6):
                chunk = values[i:i + 6]
                formatted_lines.append(
                    '    ' + ', '.join(fmt_float(v) for v in chunk) + ','
                )

            lines.append(f'const float {var_name}[{count}] = {{')
            lines.extend(formatted_lines)
            lines.append('};')
        lines.append('')
        
    lines.append('#endif')
    with open(output_path, 'w', encoding='utf-8', newline='\n') as f:
        f.write('\n'.join(lines))

    print(f'Generated: {output_path}')


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, 'Motor_Para.xlsx')
    default_output = os.path.join(
        script_dir, 'Motor_Control', 'Motor_Config', 'Src', 'Motor_Lookup_Tables.c')

    output_path = sys.argv[1] if len(sys.argv) > 1 else default_output

    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} not found")
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    tables = extract_tables(xlsx_path)

    if not tables:
        print("Error: no lookup tables extracted")
        sys.exit(1)

    print(f"\nTotal {len(tables)} tables:")
    for name in sorted(tables.keys()):
        values = tables[name]
        if values and isinstance(values[0], list):
            print(f"  {name}[{len(values)}][{len(values[0])}]")
        else:
            print(f"  {name}[{len(values)}]")

    generate_c(tables, output_path)


if __name__ == '__main__':
    main()
